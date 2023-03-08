import os
from openpyxl import load_workbook
from qrMaker import qrMaker

file_name = '单管号清单.xlsx'
wb = load_workbook(file_name)
ws = wb.active



rows = ws.max_row

for i in range(2, rows):
    spool = ws.cell(i, 1).value
    cwp = ws.cell(i,2).value
    print(spool)

    qrMaker(spool,cwp)
    i = i + 1

print('finish')